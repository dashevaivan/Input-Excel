from tkinter import *
from tkinter import ttk
from tkinter.ttk import Combobox, Treeview
from tkinter.messagebox import showinfo
import tkinter.font as font
from numpy import outer
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from ttkthemes import ThemedTk
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)


root = ThemedTk(theme='blue')
root.geometry('575x375')
root.iconbitmap("cowcow.png")
root.resizable(width=False, height=False)
root.title('Pendataan Kurban')
root.configure(background='#659ac9')
frame = Frame(root)
frame.pack()
f2 = font.Font(weight='bold')
#buat workbook baru
wb = Workbook()
ws = wb.active

num = 0
#membuat input data
def page1():
    root.geometry('600x525')
    tm1.place_forget()
    tm2.place_forget()
    tm3.place_forget()
    labelphoto.place_forget()
    labeljudul.place_forget()
    
    label_nama = ttk.Label(root, text='Masukkan Nama File : ')
    label_nama.place(x=10,y=10)
    label_nama['font'] = f2
    inp_nama = ttk.Entry(root, width=45)
    inp_nama.get()
    inp_nama.place(x=10, y=40)

    def tm_save():
        wb.save(filename=f'{inp_nama.get()}.xlsx')
        ws['A1'] = "Nama Pengurban"
        ws['B1'] = "Alamat"
        ws['C1'] = "Nomor HP"
        ws['D1'] = "Hewan"
        ws['E1'] = "Harga"
        ws['F1'] = "Request"
        ws['G1'] = "Jumlah Sapi Utuh"
        ws['H1'] = "Jumlah Sapi 1/7"
        ws['I1'] = "Jumlah Kambing/Domba"
        ws['G2'] = 0
        ws['H2'] = 0
        ws['I2'] = 0

    svb = ttk.Button(root, text='Save File', command=tm_save)
    svb.place(x=290,y=35)

    label1 = ttk.Label(root, text='Nama Pengurban : ')
    label1.place(x=10, y=70)
    
    e1 = ttk.Entry(root)
    e1.place(x=10, y = 90)
    e1.get()

    label2 = ttk.Label(root, text='Masukkan Alamat : ')
    label2.place(x=10, y= 130)

    e2 = ttk.Entry(root)
    e2.place(x=10,y=150)
    e2.get()

    label3 = ttk.Label(root, text='Masukkan Nomor HP : ')
    label3.place(x=10,y=190)

    e3 = ttk.Entry(root)
    e3.place(x=10,y=210)
    e3.get()

    label4 = ttk.Label(root, text='Pilih Jenis Hewan : ')
    label4.place(x=10,y=250)

    #comboboxnya
    n = ttk.Entry()
    hewanpilih = ttk.Combobox(root, textvariable=n)
    hewanpilih['values'] = (
        'Sapi Utuh',
        'Sapi 1/7',
        'Kambing/Domba'
    )
    hewanpilih.get()
    hewanpilih.place(x=10,y=270)
    hewanpilih.current()

    label5 = ttk.Label(root, text='Pilih Harga : ')
    label5.place(x=10,y=310)

    x = ttk.Entry()
    hargapilih = ttk.Combobox(root, textvariable=x)
    hargapilih['values'] = (
        '21.500.000',
        '3.100.000'
    )
    hargapilih.get()
    hargapilih.place(x=10,y=330)
    hargapilih.current()

    label6 = ttk.Label(root, text='Masukkan Request Pengurban : ')
    label6.place(x=10,y=370)

    e6 = ttk.Entry(root)
    e6.place(x=10,y=390)
    e6.get()

    def insert_data():
        global num
        num = num + 1
        wsnum = num + 1
        ws['A'+str(wsnum)] = e1.get()
        ws['B'+str(wsnum)] = e2.get()
        ws['C'+str(wsnum)] = e3.get()
        ws['D'+str(wsnum)] = hewanpilih.get()
        ws['E'+str(wsnum)] = hargapilih.get()
        ws['F'+str(wsnum)] = e6.get()

        if hewanpilih.get() == 'Sapi Utuh':
            ws["G2"].value += 1
        elif hewanpilih.get() == 'Sapi 1/7':
            ws["H2"].value += 1
        elif hewanpilih.get() == 'Kambing/Domba':
            ws["I2"].value += 1
        e1.delete(0, END)
        e2.delete(0, END)
        e3.delete(0, END)
        hargapilih.delete(0, END)
        hewanpilih.delete(0, END)
        e6.delete(0, END)
        wb.save(filename=f'Data Hewan Kurban.xlsx')
        wb.close()

    tmsmb = ttk.Button(root, text= "Submit", command=insert_data)
    tmsmb.place(x=10,y=430)

    def kembali():
        label1.place_forget()
        label3.place_forget()
        label2.place_forget()
        label4.place_forget()
        label5.place_forget()
        label6.place_forget()
        e1.place_forget()
        e2.place_forget()
        e3.place_forget()
        hargapilih.place_forget()
        hewanpilih.place_forget()
        e6.place_forget()
        tmsmb.place_forget()
        tm_back.place_forget()
        tm1.place(x=250,y=100, width=175)
        tm2.place(x=250,y=145,width=175)
        tm3.place(x=250, y=190, width=175)
        labelphoto.place(x=100, y=100)
        labeljudul.place(x=155, y=15)
        inp_nama.place_forget()
        label_nama.place_forget()
        svb.place_forget()
        root.geometry('575x375')

    tm_back = ttk.Button(root, text='Back', command=kembali)
    tm_back.place(x=10,y=470)


##### page 2
try:
    wb2 = load_workbook('Data Hewan Kurban.xlsx')
except FileNotFoundError:
    pass
try:
    ws2 = wb2.active
except NameError:
    pass

def page2():
    root.geometry('600x500')
    tm1.place_forget()
    tm2.place_forget()
    tm3.place_forget()
    labelphoto.place_forget()
    labeljudul.place_forget()
    

    label1 = ttk.Label(root, text='Nama Pengurban : ')
    label1.place(x=10, y=10)
    
    e1 = ttk.Entry(root)
    e1.place(x=10, y = 30)
    e1.get()

    label2 = ttk.Label(root, text='Masukkan Alamat : ')
    label2.place(x=10, y= 70)

    e2 = ttk.Entry(root)
    e2.place(x=10,y=90)
    e2.get()

    label3 = ttk.Label(root, text='Masukkan Nomor HP : ')
    label3.place(x=10,y=130)

    e3 = ttk.Entry(root)
    e3.place(x=10,y=150)
    e3.get()

    label4 = ttk.Label(root, text='Pilih Jenis Hewan : ')
    label4.place(x=10,y=190)

    #comboboxnya
    n = ttk.Entry()
    hewanpilih = ttk.Combobox(root, textvariable=n)
    hewanpilih['values'] = (
        'Sapi Utuh',
        'Sapi 1/7',
        'Kambing/Domba'
    )
    hewanpilih.get()
    hewanpilih.place(x=10,y=210)
    hewanpilih.current()

    label5 = ttk.Label(root, text='Pilih Harga : ')
    label5.place(x=10,y=250)

    x = ttk.Entry()
    hargapilih = ttk.Combobox(root, textvariable=x)
    hargapilih['values'] = (
        '21.500.000',
        '3.100.000'
    )
    hargapilih.get()
    hargapilih.place(x=10,y=270)
    hargapilih.current()

    label6 = ttk.Label(root, text='Masukkan Request Pengurban : ')
    label6.place(x=10,y=310)

    e6 = ttk.Entry(root)
    e6.place(x=10,y=330)
    e6.get()

    def insert_data():
        global ws2num
        try:
            ws2num = ws2.max_row + 1
            ws2['A'+str(ws2num)] = e1.get()
            ws2['B'+str(ws2num)] = e2.get()
            ws2['C'+str(ws2num)] = e3.get()
            ws2['D'+str(ws2num)] = hewanpilih.get()
            ws2['E'+str(ws2num)] = hargapilih.get()
            ws2['F'+str(ws2num)] = e6.get()

        except NameError:
            pass

        if hewanpilih.get() == 'Sapi Utuh':
            ws2["G2"].value += 1
        elif hewanpilih.get() == 'Sapi 1/7':
            ws2["H2"].value += 1
        elif hewanpilih.get() == 'Kambing/Domba':
            ws2["I2"].value += 1

        e1.delete(0, END)
        e2.delete(0, END)
        e3.delete(0, END)
        hargapilih.delete(0, END)
        hewanpilih.delete(0, END)
        e6.delete(0, END)
        try:
            wb2.save(filename= 'Data Hewan Kurban.xlsx')
        except NameError:
            pass
   
    tmsmb = ttk.Button(root, text= "Submit", command=insert_data)
    tmsmb.place(x=10,y=370)

    def kembali():
        label1.place_forget()
        label3.place_forget()
        label2.place_forget()
        label4.place_forget()
        label5.place_forget()
        label6.place_forget()
        e1.place_forget()
        e2.place_forget()
        e3.place_forget()
        hargapilih.place_forget()
        hewanpilih.place_forget()
        e6.place_forget()
        tmsmb.place_forget()
        tm_back.place_forget()
        tm1.place(x=250,y=100, width=175)
        tm2.place(x=250,y=145,width=175)
        tm3.place(x=250, y=190, width=175)
        labelphoto.place(x=100, y=100)
        labeljudul.place(x=155, y=15)
        root.geometry('575x375')

    tm_back = ttk.Button(root, text='Back', command=kembali)
    tm_back.place(x=10,y=410)

#data visualisasi nya di page 3
try:
    wb3 = load_workbook('Data Hewan Kurban.xlsx')
except FileNotFoundError:
    pass
try:
    ws3 = wb2.active
except NameError:
    pass

def page3():
    tm1.place_forget()
    tm2.place_forget()
    tm3.place_forget()
    labelphoto.place_forget()
    labeljudul.place_forget()


    def show_graph():
        tm_1.place_forget()
        tm_2.place_forget()
        tm_3.place_forget()
        datanya = {
            'Sapi Utuh' : ws3["G2"].value, 
            'Sapi 1/7' : ws3["H2"].value, 
            'Kambing/Domba' : ws3["I2"].value
        }
        
        hewan = list(datanya.keys())
        jumlah = list(datanya.values())
        
        fig = Figure(figsize=(6,6),dpi=90)
        ax = fig.add_subplot(111)
        ax.bar(hewan,jumlah)
        canvas1 = FigureCanvasTkAgg(fig)
        canvas1.draw()

        toolbar = NavigationToolbar2Tk(canvas1,root)
        toolbar.update()
        canvas1.get_tk_widget().pack(side=TOP,fill=BOTH)
        def kembali():
            tm_back.place_forget()
            toolbar.forget()
            canvas1.get_tk_widget().forget()
            tm_1.place(x=180,y=130, width=200)
            tm_2.place(x=180, y= 170, width=200)
            tm_3.place(x=180, y= 210, width=200)

        tm_back = ttk.Button(root, text='Back', command=kembali)
        tm_back.place(x=5,y=3)

    def kembali2():
        tm1.place(x=250,y=100, width=175)
        tm2.place(x=250,y=145,width=175)
        tm3.place(x=250, y=190, width=175)
        labelphoto.place(x=100, y=100)
        labeljudul.place(x=155, y=15)
        root.geometry('575x375')
        tm_1.place_forget()
        tm_2.place_forget()
        tm_3.place_forget()
    
    def data_nama():
        tm_1.place_forget()
        tm_2.place_forget()
        tm_3.place_forget()
        root.configure(background='white')

        nama = ws3.iter_rows(min_row=2,max_row=ws3.max_row,min_col=1,max_col=1)
        datanya2 = []
        for i in nama :
            for j in i:
                datanya2.append(j.value)
                        
        columns = ('nama', 'pengurban')

        tree = ttk.Treeview(root,columns=columns, show='headings')

        tree.heading('nama', text='Nama')
        tree.heading('pengurban', text='Pengurban')

        for data in datanya2:
            tree.insert('', END, values=data)

        def item_selected(event):
            for selected_item in tree.selection():
                item = tree.item(selected_item)
            record = item['values']
            showinfo(title='Information', message=','.join(record))

        tree.bind('<<TreeviewSelect>>', item_selected)

        tree.place(x=72, y=75)
        
        scrollbar = ttk.Scrollbar(root, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.place(x=475,y=65, height=250)

        def kembali3():
            tree.place_forget()
            scrollbar.place_forget()
            tm_back.place_forget()
            tm_1.place(x=180,y=130, width=200)
            tm_2.place(x=180, y= 170, width=200)
            tm_3.place(x=180, y= 210, width=200)
            root.configure(background='#659ac9')

        tm_back = ttk.Button(root, text='Back', command=kembali3)
        tm_back.place(x=5,y=3)

    tm_1 = ttk.Button(root, text='Lihat Nama - Nama Pengurban', command=data_nama)
    tm_1.place(x=180,y=110, width=200)
    tm_2 = ttk.Button(root, text='Lihat Grafik', command=show_graph)
    tm_2.place(x=180, y= 150, width=200)
    tm_3 = ttk.Button(root,text='Back',command=kembali2)
    tm_3.place(x=180, y= 190, width=200)



#### membuat main menu
f = font.Font(family='montserrat', size=35,weight='bold')
labeljudul = ttk.Label(root, text='Welcome')
labeljudul.place(x=155, y=15)
labeljudul['font'] = f
tm1 = ttk.Button(root, text='Create New File', command=page1)
tm1.place(x=250,y=100, width=175)
tm2 = ttk.Button(root,text='Add to Existing File', command=page2)
tm2.place(x=250,y=145,width=175)
tm3 = ttk.Button(root, text='Check Data', command=page3)
tm3.place(x=250, y=190, width=175)

photo = PhotoImage(file="cow.png")
labelphoto = ttk.Label(root, image=photo)
labelphoto.place(x=100, y=100)






root.mainloop()